import openpyxl  # so we import the necessary lib to handle the excel, others you can use like pandas
import warnings
import pandas as pd
#need push file, too
warnings.simplefilter("ignore")

def removefirst6(ws):

    ws.unmerge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
    ws.unmerge_cells(start_row=2, start_column=1, end_row=5, end_column=1) #鄉鎮
    ws.unmerge_cells(start_row=2, start_column=2, end_row=5, end_column=2) #村裡別
    ws.unmerge_cells(start_row=2, start_column=3, end_row=5, end_column=3) #投票所
    ws.unmerge_cells(start_row=2, start_column=4, end_row=2, end_column=6)
    ws.unmerge_cells(start_row=3, start_column=4, end_row=5, end_column=4) #柯文哲
    ws.unmerge_cells(start_row=3, start_column=5, end_row=5, end_column=5) #賴清德
    ws.unmerge_cells(start_row=3, start_column=6, end_row=5, end_column=6) #侯友宜
    ws.unmerge_cells(start_row=2, start_column=7, end_row=5, end_column=7)
    ws.unmerge_cells(start_row=2, start_column=8, end_row=5, end_column=8)
    ws.unmerge_cells(start_row=2, start_column=9, end_row=5, end_column=9)
    ws.unmerge_cells(start_row=2, start_column=10, end_row=5, end_column=10)
    ws.unmerge_cells(start_row=2, start_column=11, end_row=5, end_column=11)
    ws.unmerge_cells(start_row=2, start_column=12, end_row=5, end_column=12)
    ws.unmerge_cells(start_row=2, start_column=13, end_row=5, end_column=13)
    ws.unmerge_cells(start_row=2, start_column=14, end_row=5, end_column=14)

    ws.delete_rows(1,5)
    ws.cell(row=1, column=1).value = '鄉鎮'
    ws.cell(row=1, column=2).value = '村里'
    ws.cell(row=1, column=3).value = '投票所編號'
    ws.cell(row=1, column=4).value = '柯P'
    ws.cell(row=1, column=5).value = '賴清德'
    ws.cell(row=1, column=6).value = '侯友宜'
    ws.cell(row=1, column=7).value = '有效票數'
    ws.cell(row=1, column=8).value = '無效票數'
    ws.cell(row=1, column=9).value = '總投票數'
    ws.cell(row=1, column=10).value = '未投出票'
    ws.cell(row=1, column=11).value = '總發出票數'
    ws.cell(row=1, column=12).value = '剩餘票數'
    ws.cell(row=1, column=13).value = '總選舉人數'
    ws.cell(row=1, column=14).value = '投票率'
    return ws

if __name__ == '__main__' :
    workbook = openpyxl.load_workbook('data/總統-A05-4-候選人得票數一覽表-各投開票所(連江縣).xlsx')  # esnsure this script is in the same directory as your excel file

     #mysheet = workbook['Sheet1'] # the sheetname containing your data
    ws = workbook.active
    removefirst6(ws)
    workbook.save('updated_tabular.xlsx')  # esnure to save the workbook after perfoming operations
    workbook.close()

    df = pd.read_excel('updated_tabular.xlsx')
    df2 = df.assign(green=(df['賴清德']/ df['有效票數']))
    print('The city is  ')
    print(df2.sort_values(['green'],ascending=False) )




