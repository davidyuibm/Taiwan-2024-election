import openpyxl  # so we import the necessary lib to handle the excel, others you can use like pandas
import warnings
import pandas as pd
import os
from sortsub import removefirst6
warnings.simplefilter("ignore")

#myroot = '$PROJECT_DIR$/data'
#os.chdir(myroot)

for x in os.listdir('data'):
   #print('file name is ', x)
   if ('~' in x ):
	   continue
   elif ('總統' in x) :
    # loading and opening the excel workbook,
     workbook = openpyxl.load_workbook('data/'+x)  # esnsure this script is in the same directory as your excel file

     #mysheet = workbook['Sheet1'] # the sheetname containing your data
     ws = workbook.active
     removefirst6(ws)
     lastcol = ws.max_column
     for row in range(ws.max_row):
        row = row + 1
#        print('row = ', row, 'value = ', ws.cell(row=row, column=1).value )

        if row == 2:
            township = ws.cell(row=2, column=1).value
            ws.cell(row=row-1, column=lastcol+1).value = 'green'
            #print('township ', ws.cell(row=7, column=1).value)
        elif isinstance(ws.cell(row=row, column=1).value, str) and len(ws.cell(row=row, column=1).value) > 0:
            township = ws.cell(row=row, column=1).value
        else:
            ws.cell(row=row, column=1).value = township
        for col in range(ws.max_column):
            col = col + 1
            cell = ws.cell(row=row, column=col).value
        #if cell is not None:
#               print(cell,  '  ' , end = '')
#               print('\n')
    #ws.cell(row=row, column=lastcol+1).value = ws.cell(row=row, column=5).value / ws.cell(row=row, column=7).value
     workbook.save('updated_tabular.xlsx')  # esnure to save the workbook after perfoming operations
     workbook.close()

     df = pd.read_excel('updated_tabular.xlsx')
     df2 = df.assign(green=(df['賴清德']/ df['有效票數']))
     print('each city 99 ', x)

     #print(df2.sort_values(['green'],ascending=False).head(10) )
     print(df2.sort_values(['green'],ascending=False)[['鄉鎮','村里', '投票所編號', '賴清德', 'green']])

